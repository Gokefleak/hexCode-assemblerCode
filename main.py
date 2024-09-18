import openpyxl
text = open('code.txt').readlines()
parts = [3, 4, 2, 32, 2]
parts2 = [4, 4, 4, 4, 4, 4, 4, 4]
parts3 = [3, 4, 2, 12, 2]
parts4 = [4, 4, 4]
parts5 = [3, 4, 2, 16, 2]
parts6 = [4, 4, 4, 4]
start = 0
result = []
comandstart = 0

#for result
comandResult = []
result_comand = ''
numberString = 0


#exel
comands = openpyxl.load_workbook('ASM AVR Codes.xlsx')
sheet = comands.active
ComandNameDict4bait = {}
ComandNameDict2bait = {}

for row in sheet.iter_rows(values_only=True):
    key = row[8]
    value = row[0]
    ComandNameDict4bait[key] = value
for row in sheet.iter_rows(values_only=True):
    key = row[4]
    value = row[0]
    ComandNameDict2bait[key] = value
def for_found4bait(bin_code):
    for i in ComandNameDict4bait:
        if bin_code in i:
            return(ComandNameDict4bait[i])


import re


def for_found2bait(bin_code):
    binary_number = bin_code.replace(' ', '')
    for pattern in ComandNameDict2bait:
        mask = pattern
        if not isinstance(pattern, str):
            pattern = str(pattern)

        pattern = pattern.replace(' ', '')
        pattern = re.sub(r'[a-zA-Z]', '.', pattern)
        pattern_regex = f"^{pattern}$"
        if re.match(pattern_regex, binary_number):
            return(ComandNameDict2bait[mask], mask.replace(' ', ''))

    return False

def kscore(k):
    kcount = ''
    minus = False
    if k[0] == '1':
        minus = True
    if minus == True:
        for i in range(len(k)):
            if k[i] == '1':
                kcount += '0'
            else:
                kcount += '1'
        kcount = int(kcount, 2)
        kcount += 1
        kcount = bin(int(str(kcount), 10))
    else:
        kcount = k
    kcount += '0'
    kcount = int(kcount, 2)
    if minus == True:
        kcount *= -1
    else:
        kcount = '+' + str(kcount)
    return kcount
def Pscore(P):
    Pcount = hex(int(P, 2))
    return Pcount
def bscore(b):
    bcount = hex(int(b, 2))
    return bcount
def dscore(d):
    dcount = int(d, 2)
    dcount += 16
    dcount = 'r' + str(dcount)
    return dcount
def rscore(r):
    rcount = int(r, 2)
    rcount = 'r' + str(rcount)
    return rcount
def Kscore(K):
    Kcount = hex(int(K, 2))
    return Kcount
def formatcomand(stroka, elements):
    global result_comand
    global numberString
    continuenext = False
    comandstart = 0
    for i in range(len(elements)):
        if continuenext == True:
            continuenext = False
            comandstart += 4
            continue
        hex_number = ''
        bin_number = bin(int(stroka[comandstart:comandstart + elements[i]][2:4] + stroka[comandstart:comandstart + elements[i]][0:2], 16))[2:]
        NameComand = for_found4bait(bin_number)
        if NameComand is not None:
            if '*jmp k' in NameComand or '*call k' in NameComand:
                continuenext = True
                bin_number = bin(int(stroka[comandstart:comandstart + elements[i]][2:4] + stroka[comandstart:comandstart + elements[i]][0:2] + stroka[comandstart + 4:comandstart + 4 + elements[i + 1]][2:4] + stroka[comandstart + 4:comandstart + 4 + elements[i + 1]][0:2], 16))[2:]
            for_hex = str(int(bin_number + '0') % 100000000)
            hex_number = hex(int(for_hex, 2))[2:]
        if NameComand is None:
            while len(bin_number) != 16:
                bin_number = '0' + bin_number
            NameComand, values = for_found2bait(bin_number)
            P = ''
            r = ''
            d = ''
            b = ''
            k = ''
            K = ''
            s = ''
            for j in range(len(bin_number)):
                if bin_number[j] != values[j]:
                    if values[j] == 'P':
                        P += bin_number[j]
                    if values[j] == 'r':
                        r += bin_number[j]
                    if values[j] == 'd':
                        d += bin_number[j]
                    if values[j] == 'b':
                        b += bin_number[j]
                    if values[j] == 'k':
                        k += bin_number[j]
                    if values[j] == 'K':
                        K += bin_number[j]
                    if values[j] == 's':
                        s += bin_number[j]
            if P != '':
                hex_number += str(Pscore(P)) + ', '
            if d != '':
                if 'eor' in NameComand:
                    hex_number += rscore(d) + ', '
                else:
                    hex_number += dscore(d) + ', '
            if b != '':
                hex_number += str(bscore(b))
            if k != '':
                hex_number += str(kscore(k))
            if K != '':
                hex_number += Kscore(K)
            if r != '':
                hex_number += rscore(r)

        if NameComand is not None:
            if '*jmp k' in NameComand or '*call k' in NameComand:
                if NameComand == '*jmp k':
                    NameComand = 'jmp'
                if NameComand == '*call k':
                    NameComand = 'call'
                hex_stroka = hex(int(str(numberString), 10))[2:]
                result_comand += str(hex_stroka) + ':  ' + stroka[comandstart:comandstart + elements[i]][0:2] + ' ' + stroka[comandstart:comandstart + elements[i]][2:4] + ' ' + stroka[comandstart + 4:comandstart + 4 + elements[i + 1]][0:2] + ' ' + stroka[comandstart + 4:comandstart + 4 + elements[i + 1]][2:4] + '  ' + NameComand + '  0x' + hex_number + '\n'
                numberString += 4
            else:
                hex_stroka = hex(int(str(numberString), 10))[2:]
                NameComand = NameComand.split(' ')[0]
                if NameComand.split(' ')[0] == 'lsl':
                    NameComand = 'non'
                result_comand += str(hex_stroka) + ':  ' + stroka[comandstart:comandstart + elements[i]][0:2] + ' ' + stroka[comandstart:comandstart + elements[i]][2:4] + '  ' + NameComand + '  ' + str(hex_number) + '\n'
                numberString += 2
        comandstart += elements[i]
lensroka = 0
def formatstr(stroka):
    start = 0
    lensroka = int(stroka[1]+stroka[2], 16)
    if (lensroka == 16):
        for part in parts:
            if len(stroka[start:start + part])==32:
                formatcomand(stroka[start:start + part], parts2)
            start += part
    if (lensroka == 6):
        for part in parts3:
            if len(stroka[start:start + part])==12:
                formatcomand(stroka[start:start + part], parts4)
            start += part
    if (lensroka == 8):
        for part in parts5:
            if len(stroka[start:start + part])==16:
                formatcomand(stroka[start:start + part], parts6)
            start += part
for i in text:
    formatstr(i)
print(result_comand)